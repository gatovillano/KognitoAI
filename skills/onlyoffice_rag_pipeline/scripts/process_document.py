"""
Script principal del pipeline OnlyOffice RAG.
Procesa documentos de OnlyOffice y los integra con el sistema de conocimiento.
"""

import logging
import uuid
from typing import Any, Type, Optional, List, Dict
from pydantic import BaseModel, Field
from langchain_core.tools import BaseTool
from langchain_core.embeddings import Embeddings
from sqlalchemy import select, text
from core.database import SessionLocal, Document
import asyncio

from core.onlyoffice_storage import resolve_onlyoffice_file_path

logger = logging.getLogger(__name__)

# Intentar importar extractores según disponibilidad
try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    from openpyxl import load_workbook
    HAS_OPENXML = True
except ImportError:
    HAS_OPENXML = False

try:
    import PyPDF2
    HAS_PDF = True
except ImportError:
    HAS_PDF = False


class ProcessDocumentInput(BaseModel):
    document_id: str = Field(..., description="ID del documento en OnlyOffice")
    chunk_size: int = Field(128, description="Tamaño de cada chunk de texto")
    overlap: int = Field(20, description="Solapamiento entre chunks")


class ProcessOnlyOfficeDocumentTool(BaseTool):
    name: str = "process_onlyoffice_document"
    description: str = (
        "Procesa un documento de OnlyOffice, extrae su texto, genera embeddings "
        "y lo almacena en el sistema de conocimiento RAG. Úsalo cuando necesites "
        "que un documento sea searchable por el agente."
    )
    args_schema: Type[BaseModel] = ProcessDocumentInput
    account_id: str = Field(..., description="ID de cuenta del usuario")

    async def _arun(
        self, 
        document_id: str, 
        chunk_size: int = 128, 
        overlap: int = 20,
        **kwargs
    ) -> str:
        try:
            # Validar y convertir IDs
            doc_uuid = uuid.UUID(document_id)
            acc_uuid = uuid.UUID(self.account_id)
            
            async with SessionLocal() as db:
                # Obtener documento de la base de datos
                stmt = select(Document).where(
                    Document.id == doc_uuid,
                    Document.account_id == acc_uuid
                )
                result = await db.execute(stmt)
                doc = result.scalar_one_or_none()
                
                if not doc:
                    return f"Documento {document_id} no encontrado o no pertenece a la cuenta."
                
                # Extraer texto según tipo de archivo
                text_content = await self._extract_text(doc)
                
                if not text_content or not text_content.strip():
                    return f"El documento {doc.filename} está vacío o no se pudo extraer texto."
                
                # Dividir en chunks
                chunks = self._split_text(text_content, chunk_size, overlap)
                
                # Generar embeddings y almacenar
                result = await self._store_chunks(db, doc, chunks, acc_uuid)
                
                return result
                
        except Exception as e:
            logger.error(f"Error procesando documento {document_id}: {e}", exc_info=True)
            return f"Error: {str(e)}"

    async def _extract_text(self, doc: Document) -> str:
        """Extrae texto del documento según su tipo."""
        filename = doc.filename.lower()
        
        # Leer contenido del archivo
        try:
            file_path = resolve_onlyoffice_file_path(doc.file_path)
        except ValueError:
            return ""

        if not file_path.exists():
            return ""
        
        text = ""
        
        if filename.endswith('.docx') and HAS_DOCX:
            text = self._extract_docx(str(file_path))
        elif filename.endswith('.xlsx') and HAS_OPENXML:
            text = self._extract_xlsx(str(file_path))
        elif filename.endswith('.pdf') and HAS_PDF:
            text = self._extract_pdf(str(file_path))
        else:
            # Intentar lectura plana
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    text = f.read()
            except:
                pass
        
        return text

    def _extract_docx(self, file_path: str) -> str:
        """Extrae texto de archivo DOCX."""
        try:
            doc = DocxDocument(file_path)
            return "\n".join(para.text for para in doc.paragraphs)
        except Exception as e:
            logger.warning(f"Error extrayendo DOCX: {e}")
            return ""

    def _extract_xlsx(self, file_path: str) -> str:
        """Extrae texto de archivo XLSX."""
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            text_parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f"Hoja: {sheet_name}")
                for row in ws.iter_rows(values_only=True):
                    text_parts.append(" | ".join(str(cell) for cell in row if cell))
            return "\n".join(text_parts)
        except Exception as e:
            logger.warning(f"Error extrayendo XLSX: {e}")
            return ""

    def _extract_pdf(self, file_path: str) -> str:
        """Extrae texto de archivo PDF."""
        try:
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.warning(f"Error extrayendo PDF: {e}")
            return ""

    def _split_text(self, text: str, chunk_size: int, overlap: int) -> List[str]:
        """Divide texto en chunks con solapamiento."""
        words = text.split()
        if len(words) <= chunk_size:
            return [" ".join(words)]
        
        chunks = []
        for i in range(0, len(words), chunk_size - overlap):
            chunk = words[i:i + chunk_size]
            if chunk:
                chunks.append(" ".join(chunk))
        return chunks

    async def _store_chunks(
        self, 
        db, 
        doc: Document, 
        chunks: List[str],
        account_id: uuid.UUID
    ) -> str:
        """Almacena chunks usando el servicio de embeddings del sistema."""
        from core.embedding_manager import get_embedding_service
        
        embedding_service = await get_embedding_service()
        stored_count = 0
        
        for i, chunk in enumerate(chunks):
            try:
                # Generar embedding usando aembed_query
                embedding = await embedding_service.aembed_query(chunk)
                
                # Insertar en langchain_pg_embedding
                stmt = text("""
                    INSERT INTO langchain_pg_embedding 
                    (uuid, embedding, document, metadatas, account_id, workspace_id, collection_id)
                    VALUES (:uuid, :embedding, :document, :metadatas, :account_id, :workspace_id, :collection_id)
                """)
                
                await db.execute(stmt, {
                    "uuid": str(uuid.uuid4()),
                    "embedding": embedding,
                    "document": chunk,
                    "metadatas": {
                        "source": "onlyoffice",
                        "document_id": str(doc.id),
                        "filename": doc.filename,
                        "chunk_index": i,
                        "total_chunks": len(chunks)
                    },
                    "account_id": str(account_id),
                    "workspace_id": str(doc.workspace_id) if doc.workspace_id else None,
                    "collection_id": None
                })
                stored_count += 1
            except Exception as e:
                logger.error(f"Error guardando chunk {i}: {e}")
        
        await db.commit()
        return f"Documento procesado: {len(chunks)} chunks almacenados en el conocimiento."

    def _run(self, *args, **kwargs):
        raise NotImplementedError("Esta herramienta solo soporta ejecución asíncrona.")
